import Queue
import threading
import time
import nmap
import requests
from openpyxl import Workbook

exitFlag = 0
num = 0

def cut_net_mask(input_ip):
    target = input_ip.split('/')[0]
    mask = input_ip.split('/')[1]
    print target, mask
    answer = []
    target = target.split('.')
    #print target
    temp = str()
    for x in target:
        x = int(x)
        x = bin(x)
        x = str(x).lstrip('0b')
        while len(x) < 8:
            x = '0' + x
        temp = temp + x
    #print temp
    #print temp[0:int(mask)] + 'x'*(32 - int(mask))
    #print int('1'*(32 - int(mask)), 2)
    #print 2 ** (32 - int(mask)) - 1
    top = 0b0
    all = str()

    while top < int('1'*(32 - int(mask)), 2):
        y = bin(top)
        y = str(y).lstrip('0b')
        while len(y) < (32 - int(mask)):
            y = '0' + y
            #print y
        all = temp[0:int(mask)] + y
        #print temp[0:int(mask)] + y
        top += 0b1

        ip1 = all[0:8]
        ip2 = all[8:16]
        ip3 = all[16:24]
        ip4 = all[24:32]
        answer.append(str(int(ip1, 2)) + '.' + str(int(ip2, 2)) +
                      '.' + str(int(ip3, 2)) + '.' + str(int(ip4, 2)))
    return answer


def post(ip_address, port, product, version, name, cpe, extrainfo):
    post_payload = {
        'ip_address': ip_address,
        'port': port,
        'product': product,
        'version': version,
        'name': name,
        'cpe': cpe,
        'extrainfo': extrainfo,
    }
    requests.post('http://127.0.0.1/scan.php', data=post_payload)


def py_scan(target):
    global num
    print 'num = %s' % num
    num2 = 0
    print 'scanning %s' % target
    try:
        nm = nmap.PortScanner()
        nm.scan(hosts=target, ports='1-1000,1433,3306,3389,8080-8090',arguments='-A -v ',sudo=True)
        print nm.command_line()
        for port in range(1, 10000):
            try:
                s1 = nm[target]['tcp'][port]['product']
                s2 = nm[target]['tcp'][port]['version']
                s3 = nm[target]['tcp'][port]['name']
                s4 = nm[target]['tcp'][port]['cpe']
                s5 = nm[target]['tcp'][port]['extrainfo']
                print target, port, s1, s2, s3, s4, s5
                post(target, port, s1, s2, s3, s4, s5)
                ws.append([target, port, s1, s2, s3, s4, s5])
                num2 += 1
            except:
                pass
        if num2 != 0:
            num += 1
        else:
            requests.post('http://127.0.0.1/del_new.php',
                          data={'ip_address': target})
    except:
        print 'scanning %s error!' % target
    if num == 50:
        num = 0
        wb.save("/shared/%s.xlsx" % time.asctime(time.localtime(time.time())).replace(' ', '_').replace(':', '.')
                )
        print 'Save file.'


def process_data(threadName, q):
    while not exitFlag:
        queueLock.acquire()
        if not workQueue.empty():
            data = q.get()
            queueLock.release()
            print "%s processing %s" % (threadName, data) + '\n'
            py_scan(data)
        else:
            queueLock.release()
        time.sleep(1)


class myThread (threading.Thread):
    def __init__(self, threadID, name, q):
        threading.Thread.__init__(self)
        self.threadID = threadID
        self.name = name
        self.q = q

    def run(self):
        print "Starting " + self.name + '\n'
        process_data(self.name, self.q)
        print "Exiting " + self.name + '\n'


f = open('/shared/target.txt', 'r')
target_list = []
lines = f.readlines()
for line in lines:
    line = line.rstrip('\r\n')
    if '/' in line:
        for x in cut_net_mask(line):
            target_list.append(x)
    else:
        target_list.append(line)
print target_list
f.close()

wb = Workbook()
ws = wb.active
ws['A1'] = 'ip_address'
ws['B1'] = 'port'
ws['C1'] = 'product'
ws['D1'] = 'version'
ws['E1'] = 'name'
ws['F1'] = 'cpe'
ws['G1'] = 'extrainfo'

threadList = []
for i in range(0, 4):
    threadList.append("Thread-%s" % (i+1))
queueLock = threading.Lock()
workQueue = Queue.Queue()
threads = []
threadID = 1

# Create new Tread
for tName in threadList:
    thread = myThread(threadID, tName, workQueue)
    thread.start()
    threads.append(thread)
    threadID += 1

# Fill Queue
queueLock.acquire()
for word in target_list:
    workQueue.put(word)
queueLock.release()

# Wait Queue empty
while not workQueue.empty():
    pass

exitFlag = 1

# Wait all thread finished
for t in threads:
    t.join()

if num > 0:
    wb.save("/shared/%s.xlsx" % time.asctime(time.localtime(time.time())).replace(' ', '_').replace(':', '.')
            )
    print 'Save last file.'
